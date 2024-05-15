from Sharepoint import SharePoint
from openpyxl import Workbook



# get CL1_S1007_Absence_Request sharepoint list
CL1_S1007_Absence_Request = SharePoint().connect_to_list(list_name='CL1_S1007_Absence_Request')


# create excel workbook
wb = Workbook()

dest_filepath = 'CL1_S1007_Absence_Request_list.xlsx'

# create worksheet
ws = wb.active
ws.title = 'CL1_S1007_Absence_Request List'

# setting SharePoint list values to excel cells
for idx, client in enumerate(CL1_S1007_Absence_Request, 1):
    ws.cell(column=1, row=idx, value=client.get('Title', ''))
    ws.cell(column=2, row=idx, value=client.get('AddressInfo: Street', '')) 
    ws.cell(column=3, row=idx, value=client.get('AddressInfo: City', ''))


# save workbook
wb.save(filename=dest_filepath) 
